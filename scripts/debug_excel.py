import os
import sys
import logging
from pathlib import Path
import io
from dotenv import load_dotenv
import openpyxl
import msoffcrypto

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

# Load .env file
env_path = Path(__file__).parent.parent / ".env"
load_dotenv(env_path)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def inspect_excel():
    file_path = "/app/OldProjects/v2_0/main/server/all_pc.xlsx"
    password = os.getenv("EXCEL_PASS")
    
    if not password:
        logger.error("No password found")
        return

    try:
        decrypted = io.BytesIO()
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        
        wb = openpyxl.load_workbook(decrypted, read_only=True)
        sheet = wb.active
        
        print("=== EXCEL STRUCTURE DUMP ===")
        for i, row in enumerate(sheet.iter_rows(max_row=5, values_only=True)):
            print(f"\nRow {i}:")
            for idx, cell in enumerate(row):
                if cell:
                    print(f"  Col {idx} ({openpyxl.utils.get_column_letter(idx+1)}): {cell}")
                    
    except Exception as e:
        logger.error(f"Error: {e}")

if __name__ == "__main__":
    inspect_excel()

