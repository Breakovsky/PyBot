"""
Скрипт для импорта данных сотрудников из Excel файла в БД.
"""

import asyncio
import sys
import os
import io
from pathlib import Path
from urllib.parse import quote_plus
from typing import List, Dict, Optional

# Настройка кодировки для Windows
if sys.platform == "win32":
    try:
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')
    except (AttributeError, ValueError):
        pass

# Добавляем путь к модулям
sys.path.insert(0, str(Path(__file__).parent.parent / "main"))

from dotenv import load_dotenv
from database.connection import init_db_pool, close_db_pool
from config.settings import init_settings
from config.security import get_security_manager

try:
    import openpyxl
except ImportError:
    print("❌ Ошибка: Не установлен модуль openpyxl")
    print("   Установите: pip install openpyxl")
    sys.exit(1)

try:
    import msoffcrypto
    HAS_MSOFFCRYPTO = True
except ImportError:
    HAS_MSOFFCRYPTO = False
    print("⚠️  Модуль msoffcrypto не установлен. Защищенные Excel файлы не поддерживаются.")
    print("   Для поддержки защищенных файлов: pip install msoffcrypto-tool")


async def import_from_excel(excel_path: str, sheet_name: Optional[str] = None, password: Optional[str] = None):
    """
    Импортирует данные сотрудников из Excel файла.
    
    Ожидаемые колонки в Excel:
    - ФИО (или Имя)
    - Отдел (или Подразделение)
    - Рабочая станция (или WS, Компьютер)
    - IP адрес
    - Телефон
    - AD аккаунт (или Учетная запись)
    - Примечания (опционально)
    """
    
    # Загружаем .env если есть
    env_path = Path(__file__).parent.parent / ".env"
    if env_path.exists():
        load_dotenv(env_path)
    
    # Подключаемся к БД
    security = get_security_manager()
    db_password = security.get_secret("DB_PASSWORD") or os.getenv("DB_PASSWORD", "")
    
    db_host = os.getenv("DB_HOST", "localhost")
    db_port = os.getenv("DB_PORT", "5432")
    db_name = os.getenv("DB_NAME", "tbot")
    db_user = os.getenv("DB_USER", "tbot")
    
    db_user_escaped = quote_plus(db_user)
    if db_password:
        db_password_escaped = quote_plus(db_password)
        dsn = f"postgresql://{db_user_escaped}:{db_password_escaped}@{db_host}:{db_port}/{db_name}"
    else:
        dsn = f"postgresql://{db_user_escaped}@{db_host}:{db_port}/{db_name}"
    
    db_pool = init_db_pool(dsn, min_size=2, max_size=5)
    await db_pool.initialize()
    
    print("\n" + "=" * 60)
    print("Импорт данных сотрудников из Excel")
    print("=" * 60 + "\n")
    
    # Проверяем файл
    # Исправляем кодировку пути для Windows
    if sys.platform == "win32":
        try:
            # Пробуем декодировать путь как UTF-8, если он пришел в неправильной кодировке
            if isinstance(excel_path, bytes):
                excel_path = excel_path.decode('utf-8')
            elif isinstance(excel_path, str):
                # Если путь содержит неправильно закодированные символы, пробуем исправить
                try:
                    excel_path.encode('utf-8')
                except UnicodeEncodeError:
                    # Пробуем найти файл по шаблону
                    import glob
                    parent_dir = Path(excel_path).parent
                    if parent_dir.exists():
                        pattern = str(parent_dir / "*V2.0.xlsx")
                        matches = glob.glob(pattern)
                        if matches:
                            excel_path = matches[0]
                            print(f"📝 Найден файл: {excel_path}")
        except Exception as e:
            print(f"⚠️  Предупреждение при обработке пути: {e}")
    
    excel_file = Path(excel_path)
    if not excel_file.exists():
        print(f"⚠️  Файл не найден по указанному пути: {excel_path}")
        print(f"   Пробуем найти файл в директории...")
        # Пробуем найти файл в директории
        parent = excel_file.parent
        if parent.exists():
            xlsx_files = list(parent.glob("*.xlsx"))
            if xlsx_files:
                print(f"💡 Найдены Excel файлы в директории {parent}:")
                for f in xlsx_files[:5]:
                    print(f"   - {f.name}")
                
                # Если файл один, используем его
                if len(xlsx_files) == 1:
                    excel_file = xlsx_files[0]
                    excel_path = str(excel_file)
                    print(f"✅ Используем найденный файл: {excel_file.name}")
                elif excel_file.name in [f.name for f in xlsx_files]:
                    # Если имя файла совпадает, используем найденный
                    excel_file = next(f for f in xlsx_files if f.name == excel_file.name)
                    excel_path = str(excel_file)
                    print(f"✅ Используем найденный файл: {excel_file.name}")
                else:
                    print(f"❌ Не удалось автоматически определить файл")
                    print(f"   Укажите точный путь к файлу")
                    await db_pool.close()
                    return
            else:
                print(f"❌ Excel файлы не найдены в директории {parent}")
                await db_pool.close()
                return
        else:
            print(f"❌ Директория не существует: {parent}")
            await db_pool.close()
            return
    
    print(f"📄 Открываем файл: {excel_path}")
    
    # Пытаемся получить пароль из .env если не указан
    if not password:
        password = os.getenv("EXCEL_PASSWORD", "")
        if password:
            print("🔐 Пароль найден в переменных окружения")
    
    # Открываем Excel файл
    wb = None
    try:
        # Сначала пробуем открыть как обычный файл
        try:
            wb = openpyxl.load_workbook(excel_file, data_only=True)
            print("✅ Файл открыт без пароля")
        except Exception as e1:
            error_str = str(e1).lower()
            # Если ошибка "not a zip file" или "bad zipfile", значит файл защищен
            if "not a zip file" in error_str or "bad zipfile" in error_str or "file is not a zip file" in error_str:
                # Пробуем как защищенный файл
                if not HAS_MSOFFCRYPTO:
                    print("❌ Файл защищен паролем, но модуль msoffcrypto не установлен")
                    print("   Установите: pip install msoffcrypto-tool")
                    print("   Затем используйте: python scripts\\import_employees.py \"путь\" --password \"пароль\"")
                    await db_pool.close()
                    return
                
                if not password:
                    # Пробуем получить из .env
                    password = os.getenv("EXCEL_PASSWORD", "")
                    if not password:
                        # Запрашиваем пароль интерактивно
                        import getpass
                        print("🔐 Файл защищен паролем. Введите пароль:")
                        password = getpass.getpass("Пароль: ")
                
                if password:
                    import io
                    decrypted = io.BytesIO()
                    with open(excel_file, "rb") as f:
                        office_file = msoffcrypto.OfficeFile(f)
                        office_file.load_key(password=password)
                        office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    wb = openpyxl.load_workbook(decrypted, data_only=True)
                    print("✅ Файл открыт с паролем")
                else:
                    raise Exception("Пароль не указан для защищенного файла")
            else:
                # Другая ошибка - пробуем как защищенный, если указан пароль
                if password and HAS_MSOFFCRYPTO:
                    try:
                        import io
                        decrypted = io.BytesIO()
                        with open(excel_file, "rb") as f:
                            office_file = msoffcrypto.OfficeFile(f)
                            office_file.load_key(password=password)
                            office_file.decrypt(decrypted)
                        decrypted.seek(0)
                        wb = openpyxl.load_workbook(decrypted, data_only=True)
                        print("✅ Файл открыт с паролем")
                    except Exception as e2:
                        raise e1  # Возвращаем исходную ошибку
                else:
                    raise e1
        
        if not wb:
            raise Exception("Не удалось открыть файл")
        
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        print(f"📊 Лист: {ws.title}")
    except Exception as e:
        print(f"❌ Ошибка открытия файла: {e}")
        error_str = str(e).lower()
        if "password" in error_str or "encrypted" in error_str or "not a zip file" in error_str:
            print("   💡 Файл защищен паролем.")
            if not HAS_MSOFFCRYPTO:
                print("   Установите: pip install msoffcrypto-tool")
            print("   Используйте: python scripts\\import_employees.py \"путь\" --password \"пароль\"")
            print("   Или добавьте EXCEL_PASSWORD в .env файл")
        await db_pool.close()
        return
    
    # Ищем заголовки (первая строка)
    headers = []
    header_row = None
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if any(cell and isinstance(cell, str) and len(str(cell).strip()) > 0 for cell in row):
            # Пробуем найти строку с заголовками
            row_str = [str(cell).lower().strip() if cell else "" for cell in row]
            if any(keyword in " ".join(row_str) for keyword in ["фио", "имя", "отдел", "рабочая", "ip", "телефон"]):
                headers = [str(cell).strip() if cell else "" for cell in row]
                header_row = row_idx
                break
    
    if not headers:
        print("❌ Не найдена строка с заголовками")
        await db_pool.close()
        return
    
    print(f"✅ Найдены заголовки в строке {header_row}: {headers}")
    
    # Определяем индексы колонок
    col_map = {}
    for idx, header in enumerate(headers):
        if not header:
            continue
        header_lower = header.lower().strip()
        
        # ФИО - может быть в разных колонках
        if ("фио" in header_lower or "имя" in header_lower or "ф.и.о" in header_lower or 
            "user" in header_lower or "подразделение" in header_lower):
            # Проверяем, не является ли это отделом
            if "отдел" not in header_lower or "user" in header_lower:
                if "full_name" not in col_map:  # Берем первую подходящую
                    col_map["full_name"] = idx
        
        # Отдел
        if "отдел" in header_lower and "местоположение" in header_lower:
            col_map["department"] = idx
        elif "отдел" in header_lower and "department" not in col_map:
            col_map["department"] = idx
        
        # Рабочая станция
        if header_lower == "ws" or "рабочая" in header_lower or "workstation" in header_lower:
            col_map["workstation"] = idx
        
        # IP адрес (может быть в примечаниях или отдельной колонке)
        if "ip" in header_lower and "адрес" in header_lower:
            col_map["ip_address"] = idx
        
        # Телефон
        if "телефон" in header_lower or "phone" in header_lower or "тел" in header_lower:
            col_map["phone"] = idx
        
        # AD аккаунт
        if header_lower == "ad" or "аккаунт" in header_lower or "account" in header_lower:
            col_map["ad_account"] = idx
        
        # Примечания
        if "примечание" in header_lower or "заметка" in header_lower or "notes" in header_lower:
            col_map["notes"] = idx
    
    # Если не нашли ФИО, пробуем найти по позиции (обычно вторая колонка)
    if "full_name" not in col_map and len(headers) > 1:
        # Проверяем вторую колонку - часто там ФИО
        second_col = headers[1].lower().strip() if headers[1] else ""
        if second_col and ("user" in second_col or "подразделение" in second_col):
            col_map["full_name"] = 1
            print(f"💡 Используем колонку '{headers[1]}' для ФИО")
    
    if "full_name" not in col_map:
        print("❌ Не найдена колонка с ФИО")
        print(f"   Найденные колонки: {[h for h in headers if h]}")
        print("   Ищем колонки с: ФИО, Имя, USER, подразделение")
        await db_pool.close()
        return
    
    print(f"✅ Найдены колонки: {col_map}")
    
    # Читаем данные - ВСЕ строки, где есть хотя бы какая-то информация
    employees = []
    skipped_empty = 0
    skipped_no_name = 0
    
    # Определяем максимальную строку с данными
    max_row = ws.max_row
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1, max_row=max_row, values_only=True), header_row + 1):
        # Проверяем, есть ли хотя бы одно непустое значение в строке
        has_data = any(cell is not None and str(cell).strip() for cell in row if cell is not None)
        
        if not has_data:
            skipped_empty += 1
            continue
        
        emp = {}
        for key, col_idx in col_map.items():
            if col_idx < len(row):
                value = row[col_idx]
                if value is not None:
                    emp[key] = str(value).strip()
                else:
                    emp[key] = None
            else:
                emp[key] = None
        
        # Если нет ФИО, но есть другие данные - создаем запись с ФИО из других полей или используем WS/отдел
        if not emp.get("full_name") or not emp.get("full_name").strip():
            # Пробуем использовать WS или отдел как имя, если есть
            if emp.get("workstation") and emp.get("workstation").strip():
                emp["full_name"] = emp["workstation"]
            elif emp.get("department") and emp.get("department").strip():
                emp["full_name"] = emp["department"]
            else:
                # Если совсем нет данных для идентификации - пропускаем
                skipped_no_name += 1
                continue
        
        employees.append(emp)
    
    if skipped_empty > 0:
        print(f"⚠️  Пропущено {skipped_empty} полностью пустых строк")
    if skipped_no_name > 0:
        print(f"⚠️  Пропущено {skipped_no_name} строк без идентификатора (ФИО/WS/Отдел)")
    
    print(f"📋 Найдено {len(employees)} сотрудников для импорта\n")
    
    if not employees:
        print("⚠️  Нет данных для импорта")
        await db_pool.close()
        return
    
    # Импортируем в БД
    async with db_pool.acquire() as conn:
        # Получаем или создаем отделы
        departments = {}
        dept_rows = await conn.fetch("SELECT id, name FROM employees.departments")
        for row in dept_rows:
            departments[row['name'].lower()] = row['id']
        
        # Получаем или создаем рабочие станции
        workstations = {}
        ws_rows = await conn.fetch("SELECT id, name, ip_address FROM employees.workstations")
        for row in ws_rows:
            key = row['name'].lower() if row['name'] else None
            if key:
                workstations[key] = row['id']
            if row['ip_address']:
                workstations[row['ip_address']] = row['id']
        
        imported = 0
        updated = 0
        errors = 0
        
        for idx, emp in enumerate(employees, 1):
            try:
                # Получаем или создаем отдел
                dept_id = None
                if emp.get("department"):
                    dept_name = emp["department"].strip()
                    dept_key = dept_name.lower()
                    if dept_key in departments:
                        dept_id = departments[dept_key]
                    else:
                        # Создаем новый отдел
                        dept_id = await conn.fetchval("""
                            INSERT INTO employees.departments (name, description)
                            VALUES ($1, $2)
                            RETURNING id
                        """, dept_name, f"Импортирован из Excel")
                        departments[dept_key] = dept_id
                
                # Получаем или создаем рабочую станцию
                ws_id = None
                if emp.get("workstation") or emp.get("ip_address"):
                    ws_name = emp.get("workstation", "").strip() if emp.get("workstation") else None
                    ip_addr = emp.get("ip_address", "").strip() if emp.get("ip_address") else None
                    
                    # Ищем по имени или IP
                    found_ws_id = None
                    if ws_name:
                        ws_key = ws_name.lower()
                        if ws_key in workstations:
                            found_ws_id = workstations[ws_key]
                    if not found_ws_id and ip_addr:
                        if ip_addr in workstations:
                            found_ws_id = workstations[ip_addr]
                    
                    if found_ws_id:
                        ws_id = found_ws_id
                    else:
                        # Создаем новую рабочую станцию
                        ws_id = await conn.fetchval("""
                            INSERT INTO employees.workstations (name, ip_address)
                            VALUES ($1, $2)
                            RETURNING id
                        """, ws_name, ip_addr)
                        if ws_name:
                            workstations[ws_name.lower()] = ws_id
                        if ip_addr:
                            workstations[ip_addr] = ws_id
                
                # Импортируем КАЖДУЮ строку как отдельную запись
                # Проверяем, существует ли ТОЧНАЯ комбинация (ФИО + WS + Отдел)
                # Это позволяет иметь несколько записей для одного человека с разными WS
                existing = await conn.fetchrow("""
                    SELECT id FROM employees.employees
                    WHERE LOWER(full_name) = LOWER($1)
                      AND (workstation_id = $2 OR (workstation_id IS NULL AND $2 IS NULL))
                      AND (department_id = $3 OR (department_id IS NULL AND $3 IS NULL))
                    ORDER BY id
                    LIMIT 1
                """, emp["full_name"], ws_id, dept_id)
                
                if existing:
                    # Обновляем существующую запись с такой же комбинацией
                    await conn.execute("""
                        UPDATE employees.employees
                        SET department_id = $1,
                            workstation_id = $2,
                            phone = $3,
                            ad_account = $4,
                            notes = $5,
                            updated_by = $6,
                            updated_at = NOW()
                        WHERE id = $7
                    """, dept_id, ws_id, emp.get("phone"), emp.get("ad_account"), 
                        emp.get("notes"), "import_script", existing['id'])
                    updated += 1
                    if idx % 50 == 0:  # Показываем каждые 50 записей
                        print(f"🔄 Обновлен: {emp['full_name']} (запись {idx}/{len(employees)})")
                else:
                    # Создаем новую запись (даже если ФИО повторяется - это может быть другая WS/отдел)
                    await conn.execute("""
                        INSERT INTO employees.employees 
                            (full_name, department_id, workstation_id, phone, ad_account, notes, updated_by)
                        VALUES ($1, $2, $3, $4, $5, $6, $7)
                    """, emp["full_name"], dept_id, ws_id, emp.get("phone"), 
                        emp.get("ad_account"), emp.get("notes"), "import_script")
                    imported += 1
                    if idx % 50 == 0:  # Показываем каждые 50 записей
                        print(f"➕ Импортирован: {emp['full_name']} (запись {idx}/{len(employees)})")
            
            except Exception as e:
                errors += 1
                print(f"❌ Ошибка при импорте {emp.get('full_name', 'N/A')}: {e}")
    
    print("\n" + "=" * 60)
    print(f"Итоги:")
    print(f"  📥 Импортировано новых: {imported}")
    print(f"  🔄 Обновлено существующих: {updated}")
    print(f"  ❌ Ошибок: {errors}")
    print(f"  📊 Всего обработано: {imported + updated}")
    print(f"  📋 Всего строк в Excel: {len(employees)}")
    if skipped_empty > 0 or skipped_no_name > 0:
        print(f"  ⚠️  Пропущено пустых/невалидных: {skipped_empty + skipped_no_name}")
    print("=" * 60)
    
    await db_pool.close()
    print("\n✅ Импорт завершен!")


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description="Импорт данных сотрудников из Excel в БД")
    parser.add_argument("excel_path", help="Путь к Excel файлу")
    parser.add_argument("--sheet", "-s", help="Имя листа (по умолчанию активный лист)")
    parser.add_argument("--password", "-p", help="Пароль для защищенного Excel файла")
    
    args = parser.parse_args()
    
    try:
        asyncio.run(import_from_excel(args.excel_path, args.sheet, args.password))
    except KeyboardInterrupt:
        print("\n\nПрервано пользователем")
        sys.exit(1)
    except Exception as e:
        print(f"\n\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

