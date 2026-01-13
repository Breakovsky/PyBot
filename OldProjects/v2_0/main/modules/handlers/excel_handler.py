# modules/handlers/excel_handler.py

import openpyxl
import msoffcrypto
import io
import re
import logging
from tabulate import tabulate
from utils.formatters import escape_markdown_v2


def load_protected_excel(file_path: str, password: str):
    try:
        decrypted = io.BytesIO()
        with open(file_path, "rb") as f:
            office_file = msoffcrypto.OfficeFile(f)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        return openpyxl.load_workbook(decrypted, read_only=True)
    except Exception as e:
        logging.error(f"Ошибка при загрузке Excel файла: {e}")
        return None


def is_ip(value: str) -> bool:
    pattern = re.compile(r"^(?:\d{1,3}\.){3}\d{1,3}$")
    if pattern.match(value.strip()):
        return all(0 <= int(part) <= 255 for part in value.split("."))
    return False


def is_ws(value: str) -> bool:
    return value.lower().startswith('ws')


def is_digit(value: str) -> bool:
    return value.isdigit()


def _format_results_as_html_table(results):
    headers = ["ФИО", "Подразделение", "WS/WorkStation", "Телефон", "AD", "Примечание"]
    table_data = []
    for r in results:
        table_data.append([
            escape_markdown_v2(r["ФИО"]),
            escape_markdown_v2(r["Подразделение"]),
            escape_markdown_v2(r["WorkStation"]),
            escape_markdown_v2(r["Телефон"]),
            escape_markdown_v2(r["AD"]),
            escape_markdown_v2(r["Примечание"]),
        ])
    tbl = tabulate(table_data, headers=headers, tablefmt="github", colalign=("left",)*6)
    return f"<pre>{tbl}</pre>"


def _split_html_table(table_html, max_len=4000):
    if len(table_html) <= max_len:
        return table_html
    chunks = []
    current = "<pre>\n"
    lines = table_html.replace("<pre>", "").replace("</pre>", "").split("\n")
    for line in lines:
        if len(current) + len(line) + 1 > max_len - len("</pre>"):
            current += "</pre>"
            chunks.append(current)
            current = "<pre>\n" + line + "\n"
        else:
            current += line + "\n"
    current += "</pre>"
    chunks.append(current)
    return chunks


def find_in_column_g(value_g: str, wb):
    value_g = value_g.strip().lower()
    sheet = wb.active
    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[6] and str(row[6]).strip().lower() == value_g:
            results.append({
                "ФИО": str(row[1]),
                "Подразделение": str(row[2]),
                "WorkStation": str(row[7]),
                "Телефон": str(row[6]),
                "AD": str(row[14]) if len(row) > 14 else "",
                "Примечание": str(row[19]) if len(row) > 19 else ""
            })
    if not results:
        return "<pre>Значение не найдено.</pre>"
    html_table = _format_results_as_html_table(results)
    return _split_html_table(html_table)


def find_in_column_h(ws_value: str, wb):
    ws_value = ws_value.strip().lower()
    sheet = wb.active
    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[7] and ws_value in str(row[7]).strip().lower():
            results.append({
                "ФИО": str(row[1]),
                "Подразделение": str(row[2]),
                "WorkStation": str(row[7]),
                "Телефон": str(row[6]),
                "AD": str(row[14]) if len(row) > 14 else "",
                "Примечание": str(row[19]) if len(row) > 19 else ""
            })
    if not results:
        return "<pre>Значение не найдено.</pre>"
    html_table = _format_results_as_html_table(results)
    return _split_html_table(html_table)


def find_in_column_b(partial_value: str, wb):
    val = partial_value.strip().lower()
    sheet = wb.active
    results = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[1] and val in str(row[1]).lower():
            results.append({
                "ФИО": str(row[1]),
                "Подразделение": str(row[2]),
                "WorkStation": str(row[7]),
                "Телефон": str(row[6]),
                "AD": str(row[14]) if len(row) > 14 else "",
                "Примечание": str(row[19]) if len(row) > 19 else ""
            })
    if not results:
        return "<pre>Совпадений не найдено.</pre>"
    html_table = _format_results_as_html_table(results)
    return _split_html_table(html_table)


async def handle_excel_search(query: str, excel_file_path: str, excel_password: str):
    """
    Обрабатывает поисковый запрос в Excel файле.
    
    Args:
        query: Поисковый запрос
        excel_file_path: Путь к Excel файлу
        excel_password: Пароль для Excel файла
    
    Returns:
        HTML строка с результатами или сообщение об ошибке
    """
    logging.info(f"Пользовательский ввод (Excel): {query}")
    
    try:
        wb = load_protected_excel(excel_file_path, excel_password)
        if not wb:
            logging.error(f"Failed to load Excel file: {excel_file_path}")
            return "<pre>❌ Ошибка при загрузке Excel-файла. Проверьте путь и доступ к файлу.</pre>"
        
        q = query.strip().lower()
        if not q:
            wb.close()
            return "<pre>❌ Пустой запрос.</pre>"
        
        try:
            if is_ip(q) or is_digit(q):
                res = find_in_column_g(q, wb)
            elif is_ws(q):
                res = find_in_column_h(q, wb)
            else:
                res = find_in_column_b(q, wb)
        except Exception as search_error:
            logging.error(f"Error during search: {search_error}", exc_info=True)
            wb.close()
            return "<pre>❌ Ошибка при поиске в файле.</pre>"
        finally:
            wb.close()
        
        return res
        
    except FileNotFoundError:
        logging.error(f"Excel file not found: {excel_file_path}")
        return "<pre>❌ Excel-файл не найден. Проверьте путь к файлу.</pre>"
    except PermissionError:
        logging.error(f"Permission denied for Excel file: {excel_file_path}")
        return "<pre>❌ Нет доступа к Excel-файлу. Проверьте права доступа.</pre>"
    except Exception as e:
        logging.error(f"Unexpected error in handle_excel_search: {e}", exc_info=True)
        return "<pre>❌ Неожиданная ошибка при работе с Excel-файлом.</pre>"
